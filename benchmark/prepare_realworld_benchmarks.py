"""Convert DocuBench and Kleister-NDA to the repository's source-to-JSON format.

No label is used in constructing a prompt.  DocuBench image-only documents are
skipped because this text-only Qwen evaluation measures extraction, not OCR.
Kleister supplies OCR text; its official canonical values are retained as gold.
"""
from __future__ import annotations

import argparse
import csv
import json
import lzma
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DATASETS = Path("/mnt/ddn/prod-runs/interns/sunghee/datasets")
PROMPT = """Extract only information supported by the source document according to the JSON Schema.
Return exactly one JSON object and no explanation.

=== Source document ===
{source}

=== JSON Schema ===
{schema}
"""


def clipped(text: str, maximum: int) -> str:
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    if len(text) <= maximum:
        return text
    head = maximum // 2
    return text[:head] + "\n\n[... middle omitted for context limit ...]\n\n" + text[-head:]


def doc_text(path: Path) -> str | None:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return "\n\n".join(page.extract_text() or "" for page in PdfReader(path).pages)
    if suffix in {".txt", ".csv", ".xml", ".html"}:
        raw = path.read_text(encoding="utf-8", errors="replace")
        if suffix == ".csv":
            return "\n".join(" | ".join(row) for row in csv.reader(raw.splitlines()))
        if suffix in {".xml", ".html"}:
            return BeautifulSoup(raw, "html.parser").get_text("\n")
        return raw
    if suffix == ".xlsx":
        book = load_workbook(path, read_only=True, data_only=True)
        return "\n\n".join(
            f"=== Sheet: {sheet.title} ===\n" + "\n".join(
                " | ".join("" if cell is None else str(cell) for cell in row)
                for row in sheet.iter_rows(values_only=True)
            )
            for sheet in book.worksheets
        )
    if suffix == ".docx":
        with zipfile.ZipFile(path) as archive:
            xml = archive.read("word/document.xml")
        root = ElementTree.fromstring(xml)
        return "\n".join(node.text or "" for node in root.iter() if node.tag.endswith("}t"))
    return None


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def prepare_docubench(root: Path, output: Path, maximum: int) -> None:
    sources = {row["doc_id"]: row for row in json.loads((root / "sources.json").read_text())}
    rows: list[dict] = []
    skipped: dict[str, str] = {}
    for schema_path in sorted((root / "schemas").glob("*.json")):
        stem = schema_path.stem
        meta = sources[stem]
        # The user explicitly excluded receipts.  This is metadata-only,
        # pre-model filtering, not result-based selection.
        if "receipt" in meta["name"].lower():
            skipped[stem] = "receipt excluded"
            continue
        document = next((root / "documents").glob(stem + ".*"), None)
        if document is None:
            skipped[stem] = "missing document"
            continue
        text = doc_text(document)
        if not text or not text.strip():
            skipped[stem] = "no machine-readable text"
            continue
        schema = json.loads(schema_path.read_text())
        gold = json.loads((root / "labels" / f"{stem}.json").read_text())
        source = clipped(text, maximum)
        rows.append({
            "stem": f"docubench_{stem}", "dataset": "DocuBench", "source_id": stem,
            "source_name": meta["name"], "source_type": meta["ftype"], "source_pages": meta["pages"],
            "user_prompt": PROMPT.format(source=source, schema=json.dumps(schema, ensure_ascii=False)),
            "gold_json": json.dumps(gold, ensure_ascii=False), "json_schema": json.dumps(schema, ensure_ascii=False),
        })
    write_jsonl(output, rows)
    output.with_suffix(".metadata.json").write_text(json.dumps({"rows": len(rows), "skipped": skipped, "max_source_chars": maximum}, indent=2) + "\n")


def parse_kleister_target(line: str) -> dict[str, list[str]]:
    values = {"effective_date": [], "jurisdiction": [], "party": [], "term": []}
    for key, value in re.findall(r"(effective_date|jurisdiction|party|term)=([^\s]+)", line):
        values[key].append(value)
    return values


def prepare_kleister(root: Path, split: str, output: Path, maximum: int) -> None:
    expected_path = root / split / "expected.tsv"
    expected = expected_path.read_text(encoding="utf-8").splitlines() if expected_path.exists() else None
    with lzma.open(root / split / "in.tsv.xz", "rt", encoding="utf-8") as handle:
        reader = csv.DictReader(handle, delimiter="\t", fieldnames=["filename", "keys", "text_djvu", "text_tesseract", "text_textract", "text_best"])
        rows_in = list(reader)
    if expected is not None and len(rows_in) != len(expected):
        raise RuntimeError(f"{split}: input/target count mismatch")
    descriptions = {
        "effective_date": "Agreement effective date. Use ISO YYYY-MM-DD; use [] only when absent.",
        "jurisdiction": "Governing jurisdiction. Use underscore-separated canonical words, e.g. New_York; [] when absent.",
        "party": "Contracting party legal name. Use underscore-separated canonical words; [] when absent.",
        "term": "Confidentiality term. Use canonical duration such as 3_years; [] when absent.",
    }
    schema = {
        "type": "object", "additionalProperties": False,
        "properties": {key: {"type": "array", "description": descriptions[key], "items": {"type": "string"}} for key in descriptions},
        "required": ["effective_date", "jurisdiction", "party", "term"],
    }
    rows: list[dict] = []
    for index, record in enumerate(rows_in):
        source = clipped(record["text_best"], maximum)
        rows.append({
            "stem": f"kleister_{split}_{index:03d}", "dataset": "Kleister-NDA", "source_id": record["filename"],
            "user_prompt": PROMPT.format(source=source, schema=json.dumps(schema)),
            **({"gold_json": json.dumps(parse_kleister_target(expected[index]))} if expected is not None else {}),
            "json_schema": json.dumps(schema),
        })
    write_jsonl(output, rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--datasets-root", type=Path, default=DEFAULT_DATASETS)
    parser.add_argument("--output-dir", type=Path, default=ROOT / "benchmark/data/realworld")
    parser.add_argument("--max-source-chars", type=int, default=16000)
    args = parser.parse_args()
    prepare_docubench(args.datasets_root / "DocuBench", args.output_dir / "docubench_nonreceipt.jsonl", args.max_source_chars)
    for split in ("train", "dev-0", "test-A"):
        prepare_kleister(args.datasets_root / "kleister-nda", split, args.output_dir / f"kleister_nda_{split}.jsonl", args.max_source_chars)
    print(f"prepared real-world rows in {args.output_dir}")


if __name__ == "__main__":
    main()
